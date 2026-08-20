import argparse
import csv
import datetime
from collections import defaultdict
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
from firebase_db import FirebaseDB


UTC = datetime.timezone.utc


def period_dates(period, anchor):
    if period == "daily":
        return anchor, anchor
    if period == "weekly":
        start = anchor - datetime.timedelta(days=anchor.weekday())
        return start, start + datetime.timedelta(days=6)
    start = anchor.replace(day=1)
    if start.month == 12:
        next_month = start.replace(year=start.year + 1, month=1)
    else:
        next_month = start.replace(month=start.month + 1)
    return start, next_month - datetime.timedelta(days=1)


def as_aware_utc(value):
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=UTC)
    return value.astimezone(UTC)


def session_seconds_by_day(
    session, timezone, range_start, range_end, current_time=None
):
    """Split one session across local midnights and return seconds per date."""
    start = as_aware_utc(session.get("check_in"))
    if session.get("check_out") is not None:
        end = as_aware_utc(session["check_out"])
    elif session.get("status") == "open":
        end = as_aware_utc(current_time) or datetime.datetime.now(UTC)
    else:
        end = as_aware_utc(session.get("last_seen"))
    if start is None or end is None or end <= start:
        return {}

    start = start.astimezone(timezone)
    end = end.astimezone(timezone)
    period_start = datetime.datetime.combine(
        range_start, datetime.time.min, tzinfo=timezone
    )
    period_end = datetime.datetime.combine(
        range_end + datetime.timedelta(days=1),
        datetime.time.min,
        tzinfo=timezone,
    )
    start = max(start, period_start)
    end = min(end, period_end)
    if end <= start:
        return {}
    result = defaultdict(float)
    cursor = start
    while cursor < end:
        next_midnight = datetime.datetime.combine(
            cursor.date() + datetime.timedelta(days=1),
            datetime.time.min,
            tzinfo=timezone,
        )
        segment_end = min(end, next_midnight)
        if range_start <= cursor.date() <= range_end:
            result[cursor.date()] += (segment_end - cursor).total_seconds()
        cursor = segment_end
    return result


def build_report(users, sessions, period, anchor, timezone_name):
    timezone = ZoneInfo(timezone_name)
    start_date, end_date = period_dates(period, anchor)
    seconds = defaultdict(lambda: defaultdict(float))
    session_counts = defaultdict(int)

    for session in sessions:
        employee_id = session.get("employee_id")
        if not employee_id:
            continue
        pieces = session_seconds_by_day(
            session, timezone, start_date, end_date
        )
        if pieces:
            session_counts[employee_id] += 1
        for day, duration in pieces.items():
            seconds[employee_id][day] += duration

    rows = []
    number_of_days = (end_date - start_date).days + 1
    dates = [start_date + datetime.timedelta(days=i) for i in range(number_of_days)]
    for user in sorted(users, key=lambda item: item["full_name"].casefold()):
        if not user.get("active", True):
            continue
        employee_id = user["user_id"]
        daily_target = float(user["required_daily_hours"])
        working_weekdays = set(
            user.get("working_weekdays", config.DEFAULT_WORKING_WEEKDAYS)
        )
        working_days = sum(day.weekday() in working_weekdays for day in dates)
        expected_hours = daily_target * working_days
        worked_hours = sum(seconds[employee_id].values()) / 3600
        attended_days = sum(value > 0 for value in seconds[employee_id].values())
        if expected_hours == 0:
            status = "NO_SCHEDULE"
        elif worked_hours >= expected_hours:
            status = "ACHIEVED"
        else:
            status = "BELOW_REQUIRED"
        rows.append({
            "employee_id": employee_id,
            "name": user["full_name"],
            "period": period,
            "period_start": start_date.isoformat(),
            "period_end": end_date.isoformat(),
            "worked_hours": round(worked_hours, 2),
            "required_hours": round(expected_hours, 2),
            "difference_hours": round(worked_hours - expected_hours, 2),
            "attended_days": attended_days,
            "session_count": session_counts[employee_id],
            "status": status,
        })
    return start_date, end_date, rows


def export_csv(rows, period, start_date, end_date, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"attendance_{period}_{start_date}_{end_date}.csv"
    columns = [
        ("Employee ID", "employee_id"), ("Employee Name", "name"),
        ("Period", "period"), ("Period Start", "period_start"),
        ("Period End", "period_end"), ("Worked Hours", "worked_hours"),
        ("Required Hours", "required_hours"),
        ("Difference Hours", "difference_hours"),
        ("Attendance Days", "attended_days"), ("Visits", "session_count"),
        ("Status", "status"),
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=[label for label, _ in columns])
        writer.writeheader()
        for row in rows:
            writer.writerow({label: row[key] for label, key in columns})
    return path


def build_visit_rows(sessions, start_date, end_date, timezone_name):
    timezone = ZoneInfo(timezone_name)
    period_start = datetime.datetime.combine(
        start_date, datetime.time.min, tzinfo=timezone
    )
    period_end = datetime.datetime.combine(
        end_date + datetime.timedelta(days=1), datetime.time.min, tzinfo=timezone
    )
    now = datetime.datetime.now(UTC)
    visits = []
    for session in sessions:
        check_in = as_aware_utc(session.get("check_in"))
        if check_in is None:
            continue
        check_out = as_aware_utc(session.get("check_out"))
        effective_end = check_out
        if effective_end is None and session.get("status") == "open":
            effective_end = now
        if effective_end is None:
            effective_end = as_aware_utc(session.get("last_seen"))
        if effective_end is None:
            continue
        local_in = check_in.astimezone(timezone)
        local_out = effective_end.astimezone(timezone)
        overlap_start = max(local_in, period_start)
        overlap_end = min(local_out, period_end)
        if overlap_end <= overlap_start:
            continue
        visits.append({
            "employee_id": session.get("employee_id", ""),
            "name": session.get("name", ""),
            "entry": overlap_start.replace(tzinfo=None),
            "exit": None if check_out is None else overlap_end.replace(tzinfo=None),
            "duration_days": (overlap_end - overlap_start).total_seconds() / 86400,
            "status": "INSIDE" if check_out is None else "COMPLETED",
            "confidence": float(session.get("recognition_confidence", 0)),
        })
    return sorted(visits, key=lambda item: item["entry"])


def signed_hours_minutes(hours):
    total_minutes = round(abs(float(hours)) * 60)
    sign = "-" if hours < 0 else "+" if hours > 0 else ""
    return f"{sign}{total_minutes // 60}:{total_minutes % 60:02d}"


def duration_hours_minutes_seconds(days):
    total_seconds = round(float(days) * 86400)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours}:{minutes:02d}:{seconds:02d}"


def visit_history_by_employee(visits):
    history = defaultdict(list)
    for visit in visits:
        exit_text = (
            visit["exit"].strftime("%d %b %H:%M")
            if visit["exit"] is not None
            else "Still inside"
        )
        duration = duration_hours_minutes_seconds(visit["duration_days"])
        history[visit["employee_id"]].append(
            f"{visit['entry']:%d %b %H:%M} → {exit_text} ({duration})"
        )
    return history


def export_visit_csv(visits, period, start_date, end_date, output_dir):
    path = Path(output_dir) / (
        f"attendance_{period}_{start_date}_{end_date}_entry_exit.csv"
    )
    fieldnames = [
        "Employee ID", "Employee Name", "Entry Time", "Exit Time",
        "Visit Duration", "State", "Recognition Confidence",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for visit in visits:
            writer.writerow({
                "Employee ID": visit["employee_id"],
                "Employee Name": visit["name"],
                "Entry Time": visit["entry"].strftime("%Y-%m-%d %H:%M:%S"),
                "Exit Time": (
                    visit["exit"].strftime("%Y-%m-%d %H:%M:%S")
                    if visit["exit"] is not None else "Still inside"
                ),
                "Visit Duration": duration_hours_minutes_seconds(
                    visit["duration_days"]
                ),
                "State": visit["status"],
                "Recognition Confidence": round(visit["confidence"], 3),
            })
    return path


def export_excel(rows, visits, period, start_date, end_date, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"attendance_{period}_{start_date}_{end_date}.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Attendance Summary"
    dark_green = "173F35"
    green = "14805E"
    pale_green = "E2F3ED"
    pale_red = "FCE8E8"
    pale_amber = "FFF0D5"
    white = "FFFFFF"
    grey = "66736E"
    thin = Side(style="thin", color="DDE6E2")

    summary.merge_cells("A1:K1")
    summary["A1"] = "ITCO CLINIC ATTENDANCE REPORT"
    summary["A1"].font = Font(size=18, bold=True, color=white)
    summary["A1"].fill = PatternFill("solid", fgColor=dark_green)
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 34
    summary["A3"] = "Report type"
    summary["B3"] = period.title()
    summary["D3"] = "Date range"
    summary["E3"] = f"{start_date:%d %b %Y} - {end_date:%d %b %Y}"
    summary["H3"] = "Timezone"
    summary["I3"] = config.CLINIC_TIMEZONE
    summary["A4"] = "Employees"
    summary["B4"] = len(rows)
    summary["D4"] = "Achieved"
    summary["E4"] = sum(row["status"] == "ACHIEVED" for row in rows)
    summary["H4"] = "Below required"
    summary["I4"] = sum(row["status"] == "BELOW_REQUIRED" for row in rows)
    for cell in ["A3", "D3", "H3", "A4", "D4", "H4"]:
        summary[cell].font = Font(bold=True, color=grey)

    headers = [
        "Employee ID", "Employee Name", "Entry / Exit History",
        "Worked Hours", "Required Hours", "Difference (H:MM)",
        "Attendance Days", "Visits", "Status", "Period Start", "Period End",
    ]
    histories = visit_history_by_employee(visits)
    header_row = 6
    for column, label in enumerate(headers, 1):
        cell = summary.cell(header_row, column, label)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=green)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)
    for row_number, item in enumerate(rows, header_row + 1):
        values = [
            item["employee_id"], item["name"],
            "\n".join(histories[item["employee_id"]]) or "No visits",
            item["worked_hours"] / 24, item["required_hours"] / 24,
            signed_hours_minutes(item["difference_hours"]),
            item["attended_days"], item["session_count"],
            item["status"].replace("_", " "), start_date, end_date,
        ]
        for column, value in enumerate(values, 1):
            cell = summary.cell(row_number, column, value)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
        summary.cell(row_number, 3).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        summary.row_dimensions[row_number].height = max(
            24, 16 * max(1, len(histories[item["employee_id"]]))
        )
        for column in (4, 5):
            summary.cell(row_number, column).number_format = "[h]:mm"
        for column in (10, 11):
            summary.cell(row_number, column).number_format = "dd mmm yyyy"
        status_cell = summary.cell(row_number, 9)
        status_cell.font = Font(bold=True)
        status_cell.fill = PatternFill(
            "solid",
            fgColor=(
                pale_green if item["status"] == "ACHIEVED"
                else pale_amber if item["status"] == "NO_SCHEDULE"
                else pale_red
            ),
        )
    summary.freeze_panes = "A7"
    summary.auto_filter.ref = f"A6:K{max(6, 6 + len(rows))}"
    widths = [16, 24, 46, 15, 16, 18, 17, 10, 20, 15, 15]
    for index, width in enumerate(widths, 1):
        summary.column_dimensions[get_column_letter(index)].width = width
    summary.sheet_view.showGridLines = False
    summary.page_setup.orientation = "landscape"
    summary.print_title_rows = "1:6"

    details = workbook.create_sheet("Entry Exit Details")
    detail_headers = [
        "Employee ID", "Employee Name", "Entry Time", "Exit Time",
        "Visit Duration", "State", "Recognition Confidence",
    ]
    for column, label in enumerate(detail_headers, 1):
        cell = details.cell(1, column, label)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=dark_green)
        cell.alignment = Alignment(horizontal="center")
    for row_number, visit in enumerate(visits, 2):
        values = [
            visit["employee_id"], visit["name"], visit["entry"],
            visit["exit"] or "Still inside", visit["duration_days"],
            visit["status"], visit["confidence"],
        ]
        for column, value in enumerate(values, 1):
            details.cell(row_number, column, value).border = Border(bottom=thin)
        details.cell(row_number, 3).number_format = "dd mmm yyyy hh:mm"
        if visit["exit"] is not None:
            details.cell(row_number, 4).number_format = "dd mmm yyyy hh:mm"
        details.cell(row_number, 5).number_format = "[h]:mm:ss"
        details.cell(row_number, 7).number_format = "0%"
    for index, width in enumerate([16, 24, 23, 23, 17, 14, 24], 1):
        details.column_dimensions[get_column_letter(index)].width = width
    details.freeze_panes = "A2"
    details.auto_filter.ref = f"A1:G{max(1, 1 + len(visits))}"
    details.sheet_view.showGridLines = False
    workbook.save(path)
    return path


def generate(period, anchor, save_to_firestore=True, db=None):
    db = db or FirebaseDB()
    users = db.load_users()
    sessions = db.load_attendance_sessions()
    start_date, end_date, rows = build_report(
        users, sessions, period, anchor, config.CLINIC_TIMEZONE
    )
    export_csv(rows, period, start_date, end_date, config.REPORTS_DIR)
    visits = build_visit_rows(
        sessions, start_date, end_date, config.CLINIC_TIMEZONE
    )
    export_visit_csv(
        visits, period, start_date, end_date, config.REPORTS_DIR
    )
    path = export_excel(
        rows, visits, period, start_date, end_date, config.REPORTS_DIR
    )
    report_id = f"{period}_{start_date}_{end_date}"
    if save_to_firestore:
        db.save_report(report_id, {
            "period": period,
            "period_start": start_date.isoformat(),
            "period_end": end_date.isoformat(),
            "timezone": config.CLINIC_TIMEZONE,
            "generated_at": datetime.datetime.now(UTC),
            "employee_count": len(rows),
            "rows": rows,
        })
    return path, rows


def main():
    parser = argparse.ArgumentParser(description="Generate attendance reports")
    parser.add_argument("period", choices=["daily", "weekly", "monthly"])
    parser.add_argument(
        "--date",
        help="Anchor date in YYYY-MM-DD format (default: today in clinic timezone)",
    )
    parser.add_argument(
        "--local-only",
        action="store_true",
        help="Create CSV without saving the report snapshot to Firestore",
    )
    args = parser.parse_args()
    timezone = ZoneInfo(config.CLINIC_TIMEZONE)
    anchor = (
        datetime.date.fromisoformat(args.date)
        if args.date
        else datetime.datetime.now(timezone).date()
    )
    path, rows = generate(args.period, anchor, not args.local_only)
    print(f"Created {path} with {len(rows)} employee rows")


if __name__ == "__main__":
    main()
